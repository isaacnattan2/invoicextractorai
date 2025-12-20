from io import BytesIO
from typing import List

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.schemas.transaction import Transaction


def generate_excel(transactions: List[Transaction]) -> BytesIO:
    data = []
    for t in transactions:
        data.append({
            "Date": t.date,
            "Description": t.description,
            "Amount": t.amount,
            "Installment": t.installment or "",
            "Currency": t.currency,
            "Page": t.page,
            "Confidence": t.confidence,
            "Bank": t.bank
        })

    df = pd.DataFrame(data, columns=[
        "Date", "Description", "Amount", "Installment", "Currency", "Page", "Confidence", "Bank"
    ])

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Transactions", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Transactions"]

        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = header_font

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=7, max_col=7):
            for cell in row:
                cell.number_format = '0.00'

        column_widths = {
            "A": 12,
            "B": 40,
            "C": 15,
            "D": 12,
            "E": 10,
            "F": 8,
            "G": 12,
            "H": 20
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

    output.seek(0)
    return output

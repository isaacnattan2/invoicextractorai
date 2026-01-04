from io import BytesIO
from typing import List, Optional

import pandas as pd
from openpyxl.styles import Font

from app.schemas.receipt import ReceiptItem


def generate_receipt_excel(
    items: List[ReceiptItem],
    market_name: Optional[str] = None,
    cnpj: Optional[str] = None,
    address: Optional[str] = None,
    access_key: Optional[str] = None,
    issue_date: Optional[str] = None
) -> BytesIO:
    data = []
    for item in items:
        data.append({
            "Market": market_name or "",
            "CNPJ": cnpj or "",
            "Address": address or "",
            "Access Key": access_key or "",
            "Issue Date": issue_date or "",
            "Item ID": item.item_id or "",
            "Item": item.item,
            "Quantity": item.quantidade,
            "Unit Value": item.valor_unitario,
            "Total Value": item.valor_total,
            "Discount": item.desconto,
            "EAN": item.ean or ""
        })

    df = pd.DataFrame(data, columns=[
        "Market", "CNPJ", "Address", "Access Key", "Issue Date", "Item ID", "Item", "Quantity", "Unit Value", "Total Value", "Discount", "EAN"
    ])

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Receipt Items", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Receipt Items"]

        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = header_font

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=9, max_col=11):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=8, max_col=8):
            for cell in row:
                cell.number_format = '0.00'

        column_widths = {
            "A": 25,
            "B": 18,
            "C": 40,
            "D": 50,
            "E": 12,
            "F": 10,
            "G": 40,
            "H": 10,
            "I": 12,
            "J": 12,
            "K": 10,
            "L": 15
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

    output.seek(0)
    return output
